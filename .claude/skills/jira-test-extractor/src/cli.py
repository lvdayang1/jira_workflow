# -*- coding: utf-8 -*-
"""
Jira Test Extractor - Unified CLI
"""
import os
import sys
import json
import click


def get_bundle_dir():
    """Get the bundle directory for PyInstaller onefile, or the script directory."""
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        # Running as compiled onefile executable
        return sys._MEIPASS
    elif getattr(sys, 'frozen', False):
        # Running as compiled (not onefile)
        return os.path.dirname(sys.executable)
    else:
        # Running as script
        return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


# Base directory for bundled resources
BUNDLE_DIR = get_bundle_dir()

# Add src directory to path for imports
SRC_DIR = os.path.join(BUNDLE_DIR, "src")
if SRC_DIR not in sys.path:
    sys.path.insert(0, SRC_DIR)

from extractor import JiraExtractor, load_config, save_ticket_data
from attachment_reader import read_attachment
from generator import convert_to_docs


def find_config():
    """Find config file, checking multiple locations."""
    search_paths = [
        os.path.join(os.getcwd(), "config.json.local"),
        os.path.join(os.getcwd(), ".claude", "config.json.local"),
        os.path.join(os.getcwd(), "config.json"),
        os.path.join(os.getcwd(), ".claude", "config.json"),
        os.path.join(BUNDLE_DIR, "config.json.local"),
        os.path.join(BUNDLE_DIR, "config.json"),
    ]
    for path in search_paths:
        if os.path.exists(path):
            return path
    return None


def get_default_output_dir():
    """Get default output directory."""
    return os.path.join(os.getcwd(), "test_cases")


@click.group()
@click.version_option(version="1.0.0")
def cli():
    """Jira Test Extractor - Unified CLI for extracting Jira tickets and generating test case documents."""
    pass


@cli.command()
@click.argument("ticket")
@click.option("--config", "-c", "config_path", default=None, help="Path to config.json")
@click.option("--output", "-o", default=None, help="Output directory")
@click.option("--no-download", is_flag=True, help="Skip attachment download")
def extract(ticket, config_path, output, no_download):
    """Extract Jira ticket information to info.json"""
    output_dir = output if output else get_default_output_dir()
    os.makedirs(output_dir, exist_ok=True)

    if config_path is None:
        config_path = find_config()

    if config_path is None or not os.path.exists(config_path):
        click.echo(f"错误: 配置文件不存在: {config_path}", err=True)
        click.echo("请使用 --config 指定配置文件，或在当前目录/.claude/目录放置 config.json", err=True)
        sys.exit(1)

    click.echo(f"配置: {config_path}")
    click.echo(f"输出目录: {output_dir}")
    click.echo(f"正在连接 Jira...")

    config = load_config(config_path)
    extractor = JiraExtractor(config)
    extractor.connect()

    try:
        click.echo(f"正在提取 Ticket 信息：{ticket}")
        ticket_data = extractor.extract_ticket(ticket)
        click.echo(f"已提取：{ticket_data['summary']}")
        click.echo(f"状态：{ticket_data['status']}")
        click.echo(f"优先级：{ticket_data['priority']}")
        click.echo(f"描述长度：{len(ticket_data['description'])} 字符")
        click.echo(f"评论数量：{len(ticket_data['comments'])}")
        click.echo(f"附件数量：{len(ticket_data['attachments'])}")

        if not no_download and ticket_data["attachments"]:
            click.echo("\n正在下载附件...")
            ticket_data["attachments"] = extractor.download_attachments(
                ticket_data["id"],
                ticket_data["attachments"],
                output_dir
            )

        ticket_dir = save_ticket_data(ticket_data, output_dir)
        click.echo(f"\n提取完成！数据保存在: {ticket_dir}")
        click.echo("\n下一步：请让 AI 模型读取该文件夹内容生成测试用例")

    finally:
        extractor.disconnect()


@cli.command("read-attachments")
@click.argument("ticket_id")
@click.option("--config", "-c", "config_path", default=None, help="Path to config.json")
@click.option("--output", "-o", default=None, help="Output file for combined content (JSON)")
def read_attachments(ticket_id, config_path, output):
    """Download and read all attachments from a ticket"""
    ticket_id = ticket_id.upper()

    if config_path is None:
        config_path = find_config()

    if config_path is None or not os.path.exists(config_path):
        click.echo(f"错误: 配置文件不存在: {config_path}", err=True)
        sys.exit(1)

    config = load_config(config_path)

    # Determine output directory and info.json path
    output_dir = get_default_output_dir()
    info_path = os.path.join(output_dir, ticket_id, "info.json")

    if not os.path.exists(info_path):
        click.echo(f"错误: 找不到 {info_path}，请先运行 extract 命令", err=True)
        sys.exit(1)

    with open(info_path, 'r', encoding='utf-8') as f:
        info = json.load(f)

    # Download attachments if needed
    attachments_dir = os.path.join(output_dir, ticket_id, "attachments")
    os.makedirs(attachments_dir, exist_ok=True)

    needs_download = any('url' in att and 'path' not in att for att in info['attachments'])

    if needs_download:
        extractor = JiraExtractor(config)
        extractor.connect()

        click.echo(f"开始下载 {len(info['attachments'])} 个附件...")

        for att in info['attachments']:
            if 'path' in att:
                click.echo(f"  跳过(已存在): {att['filename']}")
                continue

            filename = att['filename']
            url = att['url']
            filepath = os.path.join(attachments_dir, filename)

            try:
                response = extractor.session.get(url)
                if response.ok:
                    with open(filepath, 'wb') as f:
                        f.write(response.content)
                    att['path'] = filepath
                    att['status'] = 'success'
                    click.echo(f"  下载成功: {filename}")
                else:
                    click.echo(f"  下载失败: {filename} - {response.status_code}")
            except Exception as e:
                click.echo(f"  下载错误: {filename} - {e}")

        # Save updated info.json
        with open(info_path, 'w', encoding='utf-8') as f:
            json.dump(info, f, ensure_ascii=False, indent=2)

        extractor.disconnect()
    else:
        click.echo("附件已下载，直接读取...")

    # Read all attachments
    click.echo("\n开始读取附件内容...")
    attachment_contents = []

    for att in info['attachments']:
        filename = att['filename']
        filepath = os.path.join(attachments_dir, filename)

        if not os.path.exists(filepath):
            continue

        click.echo(f"  读取: {filename}")
        content = read_attachment(filepath)

        if content and not content.startswith('[不支持'):
            attachment_contents.append({
                'filename': filename,
                'content': content
            })

    # Combine and save
    combined = {
        'ticket_id': info['id'],
        'summary': info.get('summary', ''),
        'status': info.get('status', ''),
        'priority': info.get('priority', ''),
        'description': info.get('description', ''),
        'comments': info.get('comments', []),
        'attachments_summary': [
            {'filename': a['filename'], 'content_preview': a['content'][:500] + '...' if len(a['content']) > 500 else a['content']}
            for a in attachment_contents
        ],
        'full_attachment_contents': attachment_contents
    }

    if output:
        with open(output, 'w', encoding='utf-8') as f:
            json.dump(combined, f, ensure_ascii=False, indent=2)
        click.echo(f"\n组合内容已保存到: {output}")
    else:
        output_default = os.path.join(output_dir, ticket_id, "combined_content.json")
        with open(output_default, 'w', encoding='utf-8') as f:
            json.dump(combined, f, ensure_ascii=False, indent=2)
        click.echo(f"\n组合内容已保存到: {output_default}")

    # Print combined info
    click.echo("\n" + "="*60)
    click.echo(f"Ticket: {combined['ticket_id']} - {combined['summary']}")
    click.echo("="*60)

    click.echo("\n【基本信息】")
    click.echo(f"状态: {combined['status']}")
    click.echo(f"优先级: {combined['priority']}")

    click.echo("\n【描述】")
    desc = combined['description']
    click.echo(desc[:1000] + '...' if len(desc) > 1000 else desc)

    click.echo("\n【评论】")
    if combined['comments']:
        for c in combined['comments']:
            body = c.get('body', '')[:200]
            click.echo(f"  - {c.get('author', 'Unknown')} ({c.get('time', '')}): {body}")
    else:
        click.echo("  (无评论)")

    click.echo("\n【附件列表】")
    for att in combined['attachments_summary']:
        click.echo(f"  - {att['filename']}")
        preview = att['content_preview'][:300]
        click.echo(f"    内容预览: {preview}...")

    click.echo("\n【完整附件内容】")
    for att in combined['full_attachment_contents']:
        click.echo(f"\n--- {att['filename']} ---")
        content = att['content']
        if len(content) > 2000:
            click.echo(content[:2000] + '\n... (内容已截断) ...')
        else:
            click.echo(content)


@cli.command()
@click.argument("input_file")
@click.option("--output", "-o", default=None, help="Output directory")
@click.option("--template", "-t", default=None, help="Template file for document generation")
def generate(input_file, output, template):
    """Generate test case documents from JSON"""
    try:
        files = convert_to_docs(input_file, output, template)
        click.echo(f"\n转换完成！共生成 {len(files)} 个文件:")
        for f in files:
            click.echo(f"  - {f}")
    except Exception as e:
        click.echo(f"错误: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.option("--output", "-o", default=None, help="Output config file path")
def init(output):
    """Interactively create a config.json file"""
    if output is None:
        output = os.path.join(os.getcwd(), "config.json")

    click.echo("=== Jira Test Extractor 配置初始化 ===\n")

    jira_url = click.prompt("Jira URL", default="https://your-jira-domain.com")
    username = click.prompt("Username")
    password = click.prompt("Password", hide_input=True)

    config = {
        "jira": {
            "url": jira_url,
            "username": username,
            "password": password
        },
        "output_dir": "./test_cases"
    }

    os.makedirs(os.path.dirname(output) if os.path.dirname(output) else '.', exist_ok=True)

    with open(output, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    click.echo(f"\n配置文件已保存到: {output}")


@cli.command()
@click.option("--output", "-o", default=None, help="Output template file path")
def template_cmd(output):
    """Create a default Excel template file"""
    if output is None:
        output = os.path.join(os.getcwd(), "template.xlsx")

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    os.makedirs(os.path.dirname(output) if os.path.dirname(output) else '.', exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['用例编号', '用例标题', '测试模块', '测试类型', '优先级', '前置条件', '测试步骤', '预期结果', '测试数据', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = header_alignment

    col_widths = {'A': 15, 'B': 30, 'C': 20, 'D': 12, 'E': 10, 'F': 30, 'G': 40, 'H': 40, 'I': 25, 'J': 25}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output)
    click.echo(f"默认模板已创建: {output}")


@cli.command("help")
def help_cmd():
    """Show help message with usage examples"""
    help_text = """
=== Jira Test Extractor CLI ===

用法:
    jira-workflow <command> [options]

命令:
    extract <ticket>           从 Jira Ticket 提取信息到 info.json
    read-attachments <ticket>  下载并读取 Ticket 的所有附件内容
    generate <file>            将 JSON 测试用例转换为文档格式
    init                       交互式创建 config.json
    template                   创建默认 Excel 模板
    help                       显示帮助信息

示例:

1. 提取 Jira Ticket 信息:
    jira-workflow extract PROJECT-123
    jira-workflow extract https://jira.example.com/browse/PROJECT-123
    jira-workflow extract PROJECT-123 -o ./output --config ./config.json

2. 读取附件内容:
    jira-workflow read-attachments PROJECT-123
    jira-workflow read-attachments PROJECT-123 -o combined.json

3. 生成测试用例文档:
    jira-workflow generate test_cases.json
    jira-workflow generate test_cases.json -t template.xlsx

4. 初始化配置:
    jira-workflow init -o ./config.json

5. 创建模板:
    jira-workflow template -o my_template.xlsx

配置:
    配置文件应命名为 config.json，包含以下内容:
    {
        "jira": {
            "url": "https://your-jira-domain.com",
            "username": "your_username",
            "password": "your_password"
        },
        "output_dir": "./test_cases"
    }
"""
    click.echo(help_text)


def main():
    cli()


if __name__ == "__main__":
    main()
