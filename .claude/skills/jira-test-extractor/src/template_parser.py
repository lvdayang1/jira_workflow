# -*- coding: utf-8 -*-
"""
模板解析器 - 解析 Excel 模板并生成对应的测试用例 JSON 结构
"""
import os
import json
from pathlib import Path


# 默认字段映射（当没有提供模板时使用）
DEFAULT_FIELDS = [
    {"column": "A", "field_name": "用例编号", "json_key": "id"},
    {"column": "B", "field_name": "用例标题", "json_key": "name"},
    {"column": "C", "field_name": "测试模块", "json_key": "module"},
    {"column": "D", "field_name": "测试类型", "json_key": "type"},
    {"column": "E", "field_name": "优先级", "json_key": "priority"},
    {"column": "F", "field_name": "前置条件", "json_key": "precondition"},
    {"column": "G", "field_name": "测试步骤", "json_key": "steps"},
    {"column": "H", "field_name": "预期结果", "json_key": "expected_results"},
    {"column": "I", "field_name": "测试数据", "json_key": "test_data"},
    {"column": "J", "field_name": "备注", "json_key": "remarks"},
]

# 标准字段名到 JSON Key 的映射（用于识别模板中的标准字段）
STANDARD_FIELD_MAPPING = {
    "用例编号": "id",
    "用例标题": "name",
    "测试模块": "module",
    "测试类型": "type",
    "优先级": "priority",
    "前置条件": "precondition",
    "前提条件": "precondition",
    "测试步骤": "steps",
    "预期结果": "expected_results",
    "测试数据": "test_data",
    "备注": "remarks",
    "操作": "operation",
    "输入": "input",
    "测试结果": "test_result",
}


class TemplateParser:
    """解析 Excel 模板，提取字段映射"""

    def __init__(self, template_path=None):
        """
        初始化模板解析器

        Args:
            template_path: Excel 模板文件路径。如果为 None，使用默认字段。
        """
        self.template_path = template_path
        self.fields = []
        self.headers = {}

        if template_path and os.path.exists(template_path):
            self._parse_template(template_path)
        else:
            self._load_default_fields()

    def _load_default_fields(self):
        """加载默认字段"""
        self.fields = DEFAULT_FIELDS.copy()
        for f in self.fields:
            self.headers[f["column"]] = f

    def _parse_template(self, template_path):
        """解析 Excel 模板文件"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("需要 openpyxl 库来解析 Excel 模板，请运行: pip install openpyxl")

        wb = load_workbook(template_path)
        ws = wb.active

        # 解析表头（第一行）
        self.headers = {}
        for cell in ws[1]:
            if cell.value:
                col_letter = cell.column_letter
                field_name = str(cell.value).strip()
                # 动态查找标准字段映射（支持用户自定义模板）
                json_key = STANDARD_FIELD_MAPPING.get(field_name)
                if json_key:
                    self.headers[col_letter] = {
                        "column": col_letter,
                        "field_name": field_name,
                        "json_key": json_key
                    }
                else:
                    # 未知字段，使用列号作为键（自定义字段）
                    self.headers[col_letter] = {
                        "column": col_letter,
                        "field_name": field_name,
                        "json_key": field_name  # 使用字段名作为 json_key
                    }

        # 构建字段列表
        self.fields = list(self.headers.values())

        wb.close()

    def get_field_mapping(self):
        """获取字段映射列表"""
        return self.fields

    def get_headers(self):
        """获取表头字典 {column: field_info}"""
        return self.headers

    def to_json_schema(self):
        """根据实际模板字段动态生成 JSON Schema 描述（用于 AI 生成测试用例）"""
        properties = {}
        required = []

        # 根据模板中的实际字段动态生成 properties
        for f in self.fields:
            json_key = f["json_key"]
            field_name = f["field_name"]

            # 根据字段名推断类型和描述
            if json_key in ("steps", "expected_results"):
                # 数组类型（多行文本）
                prop = {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": self._get_field_description(field_name)
                }
            elif json_key in ("test_data",):
                # 对象类型
                prop = {
                    "type": "object",
                    "description": self._get_field_description(field_name)
                }
            else:
                # 字符串类型（包括 operation, input, precondition 等）
                prop = {
                    "type": "string",
                    "description": self._get_field_description(field_name)
                }

            properties[json_key] = prop

            # 必填字段：用例编号和用例标题是必需的
            if json_key in ("id", "name"):
                required.append(json_key)

        schema = {
            "$schema": "http://json-schema.org/draft-07/schema#",
            "type": "object",
            "properties": properties,
            "required": required
        }
        return schema

    def get_template_fields_description(self):
        """根据实际模板字段动态获取字段描述（用于提示 AI）"""
        lines = ["【模板字段】（请严格按此格式生成）"]
        for f in self.fields:
            field_name = f["field_name"]
            desc = self._get_field_description(field_name)
            lines.append(f"- {field_name}：{desc}")
        return "\n".join(lines)

    def _get_field_description(self, field_name):
        """根据字段名获取描述"""
        descriptions = {
            "用例编号": "用例编号，格式：TC-XXX（如 TC-001）",
            "id": "用例编号，格式：TC-XXX（如 TC-001）",
            "用例标题": "简洁描述测试目的",
            "name": "简洁描述测试目的",
            "测试模块": "如\"文件管理-用户管理\"",
            "module": "如\"文件管理-用户管理\"",
            "测试类型": "如\"功能测试\"、\"接口测试\"、\"性能测试\"等",
            "type": "如\"功能测试\"、\"接口测试\"、\"性能测试\"等",
            "优先级": "高/中/低",
            "priority": "高/中/低",
            "前置条件": "列出测试前的准备工作",
            "前提条件": "列出测试前的准备工作",
            "precondition": "列出测试前的准备工作",
            "测试步骤": "分步骤描述操作过程，每步一行",
            "steps": "分步骤描述操作过程，每步一行",
            "预期结果": "分条描述预期输出",
            "expected_results": "分条描述预期输出",
            "测试数据": "JSON 格式的测试数据（如无则留空）",
            "test_data": "JSON 格式的测试数据（如无则留空）",
            "备注": "补充说明",
            "remarks": "补充说明",
            "操作": "描述测试的操作步骤或动作",
            "operation": "描述测试的操作步骤或动作",
            "输入": "测试输入的数据或参数",
            "input": "测试输入的数据或参数",
            "测试结果": "预期测试结果或输出",
            "test_result": "预期测试结果或输出",
        }
        return descriptions.get(field_name, "根据实际业务需求填写")

    def generate_test_cases_json_structure(self):
        """生成测试用例 JSON 的基本结构"""
        return {
            "test_cases": []
        }


def load_template(template_path=None):
    """
    加载模板的便捷函数

    Args:
        template_path: Excel 模板文件路径。如果为 None，使用默认模板。

    Returns:
        TemplateParser 实例
    """
    return TemplateParser(template_path)


if __name__ == "__main__":
    # 测试代码
    parser = TemplateParser()
    print("字段映射:")
    for f in parser.get_field_mapping():
        print(f"  {f}")

    print("\n模板字段描述:")
    print(parser.get_template_fields_description())