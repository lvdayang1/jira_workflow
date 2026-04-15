# -*- coding: utf-8 -*-
"""
创建默认测试用例模板 Excel 文件
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def create_default_template(output_path=None):
    """创建默认测试用例模板"""
    if output_path is None:
        _dir = os.path.dirname(os.path.abspath(__file__))
        output_path = os.path.join(_dir, 'templates', 'default_template.xlsx')

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 设置样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 表头
    headers = ['用例编号', '用例标题', '测试模块', '测试类型', '优先级', '前置条件', '测试步骤', '预期结果', '测试数据', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = header_alignment

    # 示例数据行（可选，留空）
    # ...

    # 设置列宽
    col_widths = {'A': 15, 'B': 30, 'C': 20, 'D': 12, 'E': 10, 'F': 30, 'G': 40, 'H': 40, 'I': 25, 'J': 25}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_path)
    print(f"默认模板已创建: {output_path}")
    return output_path


if __name__ == "__main__":
    template_path = create_default_template()
    print(f"模板路径: {template_path}")