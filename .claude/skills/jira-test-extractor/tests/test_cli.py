# -*- coding: utf-8 -*-
"""
Tests for Jira Test Extractor CLI
"""
import os
import sys
import json
import tempfile
import pytest
from click.testing import CliRunner

# Add src directory to path
_skill_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_src_dir = os.path.join(_skill_dir, "src")
if _src_dir not in sys.path:
    sys.path.insert(0, _src_dir)

from src.cli import cli, help_cmd, template_cmd, generate


class TestHelpCommand:
    """Tests for help command"""

    def test_help_displays_usage(self):
        """Test that help command shows usage information"""
        runner = CliRunner()
        result = runner.invoke(cli, ['help'])
        assert result.exit_code == 0
        assert 'jira-workflow' in result.output
        assert 'extract' in result.output
        assert 'read-attachments' in result.output
        assert 'generate' in result.output

    def test_help_shows_examples(self):
        """Test that help command shows examples"""
        runner = CliRunner()
        result = runner.invoke(cli, ['help'])
        assert '示例' in result.output or 'example' in result.output.lower()


class TestTemplateCommand:
    """Tests for template command"""

    def test_template_creates_excel_file(self):
        """Test that template command creates an Excel file"""
        runner = CliRunner()
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'test_template.xlsx')
            result = runner.invoke(cli, ['template', '-o', output_path])

            assert result.exit_code == 0, f"Exit code: {result.exit_code}, Output: {result.output}"
            assert os.path.exists(output_path)

    def test_template_creates_file_with_sheet(self):
        """Test that template creates Excel file with correct sheet name"""
        runner = CliRunner()
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'test_template.xlsx')
            result = runner.invoke(cli, ['template', '-o', output_path])

            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            assert '测试用例' in wb.sheetnames


class TestGenerateCommand:
    """Tests for generate command"""

    def test_generate_requires_input_file(self):
        """Test that generate command requires an input file"""
        runner = CliRunner()
        result = runner.invoke(cli, ['generate'])
        assert result.exit_code != 0

    def test_generate_nonexistent_file(self):
        """Test that generate command handles nonexistent file"""
        runner = CliRunner()
        result = runner.invoke(cli, ['generate', 'nonexistent.json'])
        assert result.exit_code != 0
        assert '错误' in result.output or 'Error' in result.output

    def test_generate_with_valid_json(self):
        """Test generate with valid test cases JSON"""
        runner = CliRunner()
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a sample test cases JSON
            test_cases_json = {
                "id": "TEST-123",
                "summary": "Test Ticket",
                "status": "Open",
                "priority": "High",
                "test_cases": [
                    {
                        "id": "TC-001",
                        "name": "Login Test",
                        "module": "Auth",
                        "type": "功能测试",
                        "priority": "高",
                        "precondition": "User is on login page",
                        "steps": ["Enter username", "Enter password", "Click submit"],
                        "expected_results": ["Login successful", "Redirect to dashboard"]
                    }
                ]
            }

            input_file = os.path.join(tmpdir, 'test_cases.json')
            with open(input_file, 'w', encoding='utf-8') as f:
                json.dump(test_cases_json, f, ensure_ascii=False, indent=2)

            result = runner.invoke(cli, ['generate', input_file, '-o', tmpdir])

            # Should complete (may have warnings but not error)
            assert '转换完成' in result.output or result.exit_code == 0

    def test_generate_with_json_creates_all_formats(self):
        """Test that generate creates MD, DOCX and XLSX files"""
        runner = CliRunner()
        with tempfile.TemporaryDirectory() as tmpdir:
            test_cases_json = {
                "id": "TEST-123",
                "summary": "Test Ticket",
                "status": "Open",
                "priority": "High",
                "test_cases": [
                    {
                        "id": "TC-001",
                        "name": "Login Test",
                        "module": "Auth",
                        "type": "功能测试",
                        "priority": "高",
                        "precondition": "User is on login page",
                        "steps": ["Enter username", "Enter password", "Click submit"],
                        "expected_results": ["Login successful"]
                    }
                ]
            }

            input_file = os.path.join(tmpdir, 'test_cases.json')
            with open(input_file, 'w', encoding='utf-8') as f:
                json.dump(test_cases_json, f, ensure_ascii=False, indent=2)

            result = runner.invoke(cli, ['generate', input_file, '-o', tmpdir])

            # Check that files were created
            assert os.path.exists(os.path.join(tmpdir, 'test_cases.md'))
            assert os.path.exists(os.path.join(tmpdir, 'test_cases.docx'))
            assert os.path.exists(os.path.join(tmpdir, 'test_cases.xlsx'))


class TestCliVersion:
    """Tests for CLI version"""

    def test_version_option(self):
        """Test that --version works"""
        runner = CliRunner()
        result = runner.invoke(cli, ['--version'])
        assert result.exit_code == 0
        assert '1.0.0' in result.output


class TestCliHelp:
    """Tests for CLI --help option"""

    def test_help_option(self):
        """Test that --help works"""
        runner = CliRunner()
        result = runner.invoke(cli, ['--help'])
        assert result.exit_code == 0
        assert 'Commands:' in result.output


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
